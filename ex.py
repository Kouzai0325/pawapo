from openpyxl import Workbook
wb = Workbook()

ws = wb.active
ws['A1'] = 42
ws.append([1,2,3])

import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")

ab = Workbook()
colC = ws['A']
col_range = ws['A:J']
for x in range(1,101):
    for y in range(1,101):
ws.cell(row=x, column=y)
wb.save("s4mple.xlsx")

